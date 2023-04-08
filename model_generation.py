import logging
logging.getLogger("tensorflow").setLevel(logging.ERROR)
import tensorflow as tf
import tensorboard as tb
from datetime import datetime
import numpy as np
import time
import random
import pandas as pd
import openpyxl
import git
from layer_parser import parse_layers

from matplotlib import pyplot as plt
from tensorflow import keras
from keras import layers
from keras.models import Sequential
from keras.callbacks import CSVLogger

img_height = 100
img_width = 100
batch_size = 64

train_img_path = "img/dataset/train"
test_img_path = "img/dataset/test"
excel_path = "AI data recover.xlsx"


class MetricsLogger(keras.callbacks.Callback):
    def __init__(self):
        super().__init__()
        self.train_start_time = 0.0
        self.metrics = {
            'train_loss': [],
            'train_accuracy': [],
            'val_loss': [],
            'val_accuracy': [],
            'train_time': 0
        }

    def on_train_begin(self, logs=None):
        self.train_start_time = time.time()

    def on_train_end(self, logs=None):
        self.metrics['train_time'] = int(time.time() - self.train_start_time)

    def on_epoch_end(self, epoch, logs=None):
        self.metrics['train_loss'].append(logs.get('loss'))
        self.metrics['train_accuracy'].append(logs.get('accuracy'))
        self.metrics['val_loss'].append(logs.get('val_loss'))
        self.metrics['val_accuracy'].append(logs.get('val_accuracy'))

    def get_metrics(self):
        return self.metrics

def init_excel():
    new_run_number = 0
    i = 1
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.worksheets[0]
    while True:
        val = worksheet.cell(row=i,column=1).value
        if val is None:
            break
        new_run_number+=1
        i+=1
    return new_run_number

def write_to_excel(run_number, train_acc, train_loss, val_acc, val_loss, eval_acc, eval_loss, train_time, description):
    val_list = [run_number,train_acc,train_loss,val_acc,val_loss,eval_acc,eval_loss,train_time, description]
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.worksheets[0]
    for i, val in enumerate(val_list,start=1):
        worksheet.cell(row = run_number+1,column=i).value = val

    workbook.save("AI data recover.xlsx")


# take 80% of the images and use them for training
def init_datasets():
    random.seed(time.time())
    seed = random.randint(0, 999999)
    (train_ds, val_ds) = tf.keras.utils.image_dataset_from_directory(
        train_img_path,
        validation_split=0.2,
        subset="both",
        seed=seed,
        image_size=(img_height, img_width),
        batch_size=batch_size
    )

    test_ds = tf.keras.utils.image_dataset_from_directory(
        test_img_path,
        seed=seed,
        image_size=(img_height,img_width),
        batch_size=batch_size
    )
    return (train_ds, val_ds, test_ds)


def augment_image(img):
    img = tf.image.random_saturation(img, 0.9, 1.2)
    img = tf.image.random_hue(img, 0.02)
    return img

def network(num_classes):
    data_augmentation = keras.Sequential(
        [
            layers.RandomFlip("horizontal",
                              input_shape=(img_height,
                                           img_width,
                                           3)),
            layers.RandomRotation(0.1),
            layers.RandomZoom(0.1)
        ]
    )
    
    filter_size = (3, 3)
    # build the neural network model
    model = Sequential([
        data_augmentation,
        layers.Rescaling(1. / 255, input_shape=(img_height, img_width, 3)),
        layers.Conv2D(8, filter_size, strides=(1, 1), padding='same', activation='relu'),
        layers.MaxPooling2D((2, 2), strides=(2, 2), padding='valid'),
        layers.Conv2D(32, filter_size, strides=(1, 1), padding='same', activation='relu'),
        layers.MaxPooling2D((2, 2), strides=(2, 2), padding='valid'),
        layers.Conv2D(64, filter_size, strides=(1, 1), padding='same', activation='relu'),
        layers.MaxPooling2D((2, 2), strides=(2, 2), padding='valid'),
        layers.Conv2D(128, filter_size, strides=(1, 1), padding='same', activation='relu'),
        layers.MaxPooling2D((2, 2), strides=(2, 2), padding='valid'),
        layers.Flatten(),
        layers.Dense(2048, activation="relu"),
        layers.Dropout(0.2),
        layers.Dense(256, activation="relu"),
        layers.Dropout(0.2),
        layers.Dense(num_classes, activation="softmax")
    ])
    return model

def train(epochs, pre_model=None, datasets=None):
    run_num = init_excel()
    print(f"run number: {run_num}")
    if datasets is None:
        (train_ds,val_ds, test_ds) = init_datasets()
    else:
        (train_ds,val_ds, test_ds) = datasets

    class_names = train_ds.class_names
    #print(class_names)

    # buffered prefetching, so you can yield data from disk without having I/O become blocking
    AUTOTUNE = tf.data.AUTOTUNE

    train_ds = train_ds.cache().shuffle(1000).prefetch(buffer_size=AUTOTUNE)
    val_ds = val_ds.cache().prefetch(buffer_size=AUTOTUNE)

    normalization_layer = layers.Rescaling(1./255)

    normalized_ds = train_ds.map(lambda x, y: (normalization_layer(x), y))
    image_batch, labels_batch = next(iter(normalized_ds))
    first_image = image_batch[0]

    num_classes = len(class_names)

    if pre_model is None:
        model = network(num_classes)
    else:
        model = pre_model

    # compile the model
    model.compile(optimizer='adam',
                  loss=tf.metrics.sparse_categorical_crossentropy,
                  metrics=['accuracy'])

    model.summary()


    log_dir = f"boards/run {run_num}"
    tensorboard_callback=keras.callbacks.TensorBoard(log_dir=log_dir,histogram_freq=1)

    custom_metrics = MetricsLogger()
    # start training the neural network
    csv_logger = CSVLogger('training.log', separator=',', append=False)
    history = model.fit(
      train_ds,
      validation_data=val_ds,
      epochs=epochs,
      callbacks=[csv_logger,
                 tensorboard_callback,
                 custom_metrics]
    )

    print("Evaluate")
    result = model.evaluate(test_ds)
    eval_res = dict(zip(model.metrics_names, result))
    metrics = custom_metrics.get_metrics()

    model.save(f'models/run{run_num}.h5')
    write_to_excel(run_num,
                   float(round(metrics.get("train_accuracy")[-1],4)),
                   float(round(metrics.get("train_loss")[-1],4)),
                   float(round(metrics.get("val_accuracy")[-1],4)),
                   float(round(metrics.get("val_loss")[-1],4)),
                   float(round(eval_res.get("accuracy"),4)),
                   float(round(eval_res.get("loss"),4)),
                   metrics.get("train_time")//epochs,
                   parse_layers(model.layers))

    repo = git.Repo("~/AINOOBS2/AI-Noobs-Project")
    repo.git.checkout("runs")
    repo.git.add(all=True)
    repo.index.commit(f"Run #{run_num}")
    origin = repo.remote(name = "origin")
    origin.push()


def plot_data(epochs):
    log_data = pd.read_csv('training.log', sep=',', engine='python')

    acc = log_data['accuracy']
    val_acc = log_data['val_accuracy']

    loss = log_data['loss']
    val_loss = log_data['val_loss']

    epochs_range = range(epochs)

    plt.figure(figsize=(8, 8))
    plt.subplot(1, 2, 1)
    plt.plot(epochs_range, acc, label='Training Accuracy')
    plt.plot(epochs_range, val_acc, label='Validation Accuracy')
    plt.legend(loc='lower right')
    plt.title('Training and Validation Accuracy')

    plt.subplot(1, 2, 2)
    plt.plot(epochs_range, loss, label='Training Loss')
    plt.plot(epochs_range, val_loss, label='Validation Loss')
    plt.legend(loc='upper right')
    plt.title('Training and Validation Loss')
    plt.show()

def resume_training(epochs, last_epoch):
    (train_ds, val_ds) = init_datasets()
    checkpoint_dir = "models/checkpoints_RGB_1"
    log_dir="boards/resumed run"

    model = keras.models.load_model(f"models/checkpoints_RGB_1/model_epoch_{last_epoch:02d}.h5")

    csv_logger = CSVLogger('training1.log', separator=',', append=True)
    tensorboard_callback=keras.callbacks.TensorBoard(log_dir=log_dir,histogram_freq=1)

    model.fit(
        train_ds,
        validation_data=val_ds,
        epochs=epochs,
        initial_epoch=last_epoch,
        callbacks=[csv_logger, tensorboard_callback]
    )
